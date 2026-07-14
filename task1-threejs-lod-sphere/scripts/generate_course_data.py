from __future__ import annotations

import ast
import json
import math
import re
import runpy
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "src" / "data"
OUT_FILE = DATA_DIR / "course-data.json"
RADIUS = 10.0

COURSE_META = [
    ("高等数学上", "gaoshu-shang", "#2878d0"),
    ("高等数学下", "gaoshu-xia", "#d94f5c"),
    ("线性代数", "linear-algebra", "#17956f"),
    ("概率论与数理统计", "probability-statistics", "#d28a1e"),
]


def clean(value: object) -> str:
    return " ".join(str(value).strip().split()) if value is not None else ""


def normalize_xyz(xyz: object) -> np.ndarray:
    vector = np.asarray(xyz, dtype=float).reshape(3)
    norm = float(np.linalg.norm(vector))
    if norm < 1e-9:
        return np.array([0.0, 0.0, RADIUS], dtype=float)
    return vector / norm * RADIUS


def coordinate_record(xyz: object) -> dict:
    x, y, z = normalize_xyz(xyz)
    three_x, three_y, three_z = x, z, y
    phi = math.atan2(three_z, three_x)
    psi = math.asin(max(-1.0, min(1.0, three_y / RADIUS)))
    return {
        "x": round(float(x), 5),
        "y": round(float(y), 5),
        "z": round(float(z), 5),
        "phi": round(phi, 6),
        "psi": round(psi, 6),
    }


def linear_algebra_fallbacks() -> dict[str, object]:
    for path in DATA_DIR.glob("*.py"):
        source = path.read_text(encoding="utf-8")
        if "def create_all_arcs" not in source or "center_big" not in source:
            continue
        namespace = runpy.run_path(str(path))
        arcs = namespace["create_all_arcs"]()
        return namespace["collect_all_label_points"](arcs)
    return {}


def close_vertex_ids(vertex_ids: list[str]) -> list[str]:
    cleaned: list[str] = []
    for vertex_id in vertex_ids:
        if not cleaned or cleaned[-1] != vertex_id:
            cleaned.append(vertex_id)
    if len(cleaned) > 1 and cleaned[0] == cleaned[-1]:
        cleaned.pop()
    return cleaned


def parse_pair(value: object) -> tuple[str, str] | None:
    try:
        pair = ast.literal_eval(clean(value))
    except (SyntaxError, ValueError):
        return None
    if not isinstance(pair, (tuple, list)) or len(pair) != 2:
        return None
    return clean(pair[0]), clean(pair[1])


def leading_number(value: object) -> str | None:
    match = re.match(r"\s*(\d+)", clean(value))
    return match.group(1) if match else None


def build_data() -> dict:
    workbook_path = next(DATA_DIR.glob("*.xlsx"))
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    section_sheet, coordinate_sheet, pair_sheet, chapter_sheet = workbook.worksheets[:4]

    meta_by_title = {title: (course_id, color) for title, course_id, color in COURSE_META}
    records: dict[str, dict[str, dict]] = defaultdict(dict)
    name_lookup: dict[str, dict[str, str]] = defaultdict(dict)
    corrections: list[str] = []

    def add_vertex(course_title: str, point_name: str, xyz: object, source: str) -> None:
        course_id = meta_by_title[course_title][0]
        normalized_name = clean(point_name)
        key = normalized_name.casefold()
        if key in name_lookup[course_title]:
            return
        record = coordinate_record(xyz)
        record.update(
            {
                "id": f"{course_id}-vertex-{len(records[course_title]) + 1}",
                "label": normalized_name,
                "source": source,
            }
        )
        records[course_title][normalized_name] = record
        name_lookup[course_title][key] = normalized_name

    for row in coordinate_sheet.iter_rows(min_row=2, values_only=True):
        course_title, point_name, x, y, z = row[:5]
        course_title = clean(course_title)
        point_name = clean(point_name)
        if course_title not in meta_by_title or not point_name:
            continue
        if x is None or y is None or z is None:
            continue
        add_vertex(course_title, point_name, (x, y, z), "xlsx")

    for point_name, xyz in linear_algebra_fallbacks().items():
        add_vertex("线性代数", point_name, xyz, "source-fallback")

    def resolve_vertex(course_title: str, point_name: object) -> str:
        requested = clean(point_name)
        canonical = name_lookup[course_title].get(requested.casefold())
        if canonical is None:
            raise KeyError(f"Missing coordinate: {course_title} / {requested}")
        if canonical != requested:
            corrections.append(f"{course_title}: {requested} -> {canonical}")
        return records[course_title][canonical]["id"]

    chapters_by_course: dict[str, list[dict]] = defaultdict(list)
    chapter_by_title: dict[str, dict[str, dict]] = defaultdict(dict)
    chapter_by_number: dict[str, dict[str, dict]] = defaultdict(dict)
    for row in chapter_sheet.iter_rows(min_row=2, values_only=True):
        original_id, course_title, chapter_title, range_title = row[:4]
        course_title = clean(course_title)
        chapter_title = clean(chapter_title)
        if course_title not in meta_by_title or not chapter_title:
            continue
        course_id = meta_by_title[course_title][0]
        vertex_ids = close_vertex_ids(
            [resolve_vertex(course_title, value) for value in row[4:] if clean(value)]
        )
        if len(vertex_ids) < 3:
            raise ValueError(f"Chapter has fewer than three vertices: {chapter_title}")
        chapter_index = len(chapters_by_course[course_title])
        chapter_number = leading_number(range_title) or str(chapter_index)
        chapter = {
            "id": f"{course_id}-chapter-{chapter_index + 1}",
            "key": chapter_number,
            "title": chapter_title,
            "rangeTitle": clean(range_title),
            "sourceId": clean(original_id),
            "vertexIds": vertex_ids,
            "sectionIds": [],
        }
        chapters_by_course[course_title].append(chapter)
        chapter_by_title[course_title].setdefault(chapter_title, chapter)
        chapter_by_number[course_title][chapter_number] = chapter

    sections_by_course: dict[str, list[dict]] = defaultdict(list)
    for row in section_sheet.iter_rows(min_row=2, values_only=True):
        original_id, course_title, chapter_title, section_title = row[:4]
        course_title = clean(course_title)
        chapter_title = clean(chapter_title)
        section_title = clean(section_title)
        if course_title not in meta_by_title or not section_title:
            continue
        section_chapter_number = leading_number(section_title)
        chapter = None
        if section_chapter_number:
            chapter = chapter_by_number[course_title].get(section_chapter_number)
        if chapter is None:
            chapter = chapter_by_title[course_title].get(chapter_title)
        if chapter is None:
            raise KeyError(f"Missing chapter for section: {course_title} / {chapter_title} / {section_title}")
        if chapter["title"] != chapter_title:
            source_title = chapter["title"]
            corrections.append(
                f"{course_title}: chapter range '{chapter['rangeTitle']}' title '{source_title}' -> '{chapter_title}'"
            )
            chapter["sourceTitle"] = source_title
            chapter["title"] = chapter_title
        course_id = meta_by_title[course_title][0]
        section_index = len(sections_by_course[course_title])
        vertex_ids = close_vertex_ids(
            [resolve_vertex(course_title, value) for value in row[4:] if clean(value)]
        )
        if len(vertex_ids) < 3:
            raise ValueError(f"Section has fewer than three vertices: {section_title}")
        section = {
            "id": f"{course_id}-section-{section_index + 1}",
            "sourceId": clean(original_id),
            "title": section_title,
            "chapterId": chapter["id"],
            "chapterKey": chapter["key"],
            "chapter": chapter["title"],
            "vertexIds": vertex_ids,
            "knowledge": [],
        }
        sections_by_course[course_title].append(section)
        chapter["sectionIds"].append(section["id"])

    pairs_by_course: dict[str, list[dict]] = defaultdict(list)
    for row in pair_sheet.iter_rows(min_row=2, values_only=True):
        course_title, pair_value, center_name = row[:3]
        course_title = clean(course_title)
        if course_title not in meta_by_title:
            continue
        pair = parse_pair(pair_value)
        if pair is None:
            continue
        try:
            a_id = resolve_vertex(course_title, pair[0])
            b_id = resolve_vertex(course_title, pair[1])
        except KeyError:
            continue
        item = {
            "id": f"{meta_by_title[course_title][0]}-edge-{len(pairs_by_course[course_title]) + 1}",
            "a": a_id,
            "b": b_id,
        }
        center_name = clean(center_name)
        if center_name:
            item["centerName"] = center_name
            canonical_center = name_lookup[course_title].get(center_name.casefold())
            if canonical_center:
                item["center"] = records[course_title][canonical_center]["id"]
        pairs_by_course[course_title].append(item)

    courses = []
    for course_title, course_id, color in COURSE_META:
        vertices = list(records[course_title].values())
        courses.append(
            {
                "id": course_id,
                "title": course_title,
                "color": color,
                "vertices": vertices,
                "points": vertices,
                "edges": pairs_by_course[course_title],
                "chapters": chapters_by_course[course_title],
                "sections": sections_by_course[course_title],
            }
        )

    return {
        "schema": 3,
        "radius": RADIUS,
        "generatedFrom": [workbook_path.name, "线性代数各段变量名.py (P34 fallback)"],
        "corrections": sorted(set(corrections)),
        "courses": courses,
    }


def main() -> None:
    data = build_data()
    OUT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    for course in data["courses"]:
        print(
            f"{course['title']}: {len(course['vertices'])} vertices, "
            f"{len(course['chapters'])} chapters, {len(course['sections'])} sections, "
            f"{len(course['edges'])} edges"
        )
    print(f"Corrections: {len(data['corrections'])}")
    print(f"Wrote {OUT_FILE}")


if __name__ == "__main__":
    main()
